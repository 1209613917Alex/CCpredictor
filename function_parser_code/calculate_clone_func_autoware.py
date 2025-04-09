import xml.etree.ElementTree as ET
import pandas as pd
import os
import chardet
from tree_sitter_parser.tree_sitter_parser import TreeSitterParser
import glob
import openpyxl

def func_parse(codes):
    """解析代码提取函数信息"""
    tree_sitter_parser = TreeSitterParser()
    tree_sitter_parser.functions = []  # 清空已存在的函数列表
    tree_sitter_parser.tree_parser(codes)  # 解析代码
    return tree_sitter_parser.functions  # 返回解析出的函数信息


def read_cpp(cpp_adr):
    """读取cpp文件，返回读取结果，如果文件无法打开则返回 None"""
    try:
        # 首先以二进制模式读取文件
        with open(cpp_adr, 'rb') as f:
            data = f.read()

        # 检测文件的编码格式
        detected = chardet.detect(data)
        encode_str = detected['encoding']

        # 读取文件内容，尝试使用检测到的编码
        try:
            with open(cpp_adr, 'r', encoding=encode_str) as f:
                codes = f.read()
            return codes
        except (UnicodeDecodeError, TypeError):
            # 如果解码失败，尝试使用 utf-8 和 latin1
            try:
                with open(cpp_adr, 'r', encoding='utf-8') as f:
                    codes = f.read()
                return codes
            except UnicodeDecodeError:
                # 如果 utf-8 也失败，最后尝试使用 latin1
                with open(cpp_adr, 'r', encoding='latin1') as f:
                    codes = f.read()
                return codes
    except Exception as e:
        print(f"Error opening file {cpp_adr}: {e}")
        return None  # 如果文件无法打开，返回 None

def file_func_num(sourceFile):
    """获取文件中函数的数量"""
    codes = read_cpp(sourceFile)  # 读取cpp文件，获取代码进行函数解析
    functions = func_parse(codes)  # 解析cpp文件中的所有函数信息
    return len(functions)

def calculate(xml_file_path):
    # 解析 XML 文件
    tree = ET.parse(xml_file_path)
    root = tree.getroot()
    # 存储提取的数据
    data = []
    unique_files = set()

    # 遍历所有的 'dup' 节点
    for dup in root.findall('.//dup'):
        count = dup.get('count')  # 获取 'count' 属性值
        # 遍历 'dup' 下的所有 'source' 节点
        for source in dup.findall('source'):
            # 提取每个 'source' 节点的属性
            source_file = source.get('sourceFile')
            unique_files.add(source_file)
            # 提取 function_name 属性
            function_name = source.find('code').get('function_name') if source.find('code') is not None else ""
            # 将提取的信息保存到 data 列表
            data.append({
                'sourceFile': source_file,
                'function_name': function_name
            })

    # 统计每个文件的函数信息
    file_func_data = []
    for unique_file in unique_files:
        unique_func = set()
        file_func_count = file_func_num(unique_file)  # 获取文件中的函数总数
        for entry in data:
            if unique_file == entry['sourceFile']:
                unique_func.add(entry['function_name'])  # 添加克隆函数
        # 存储每个文件的函数数量和克隆函数数量
        file_func_data.append({
            'file_name': unique_file,
            'clone_func_count': len(unique_func),
            'func_count': file_func_count
        })

    # 汇总信息的字典
    results = {}

    # 计算每个文件的克隆函数对的占比（ratio）
    for file in file_func_data:
        file_name = file['file_name']
        clone_count = file['clone_func_count']
        func_count = file['func_count']
        ratio = round(clone_count / func_count,2) if func_count > 0 else 0
        # 将结果添加到汇总字典
        results[file_name] = {
            'File': file_name,
            'Clone Func Count': clone_count,
            'Func Count': func_count,
            'Clone Func Ratio': ratio
        }
    for file_name, result in results.items():
        print(f"文件名: {result['File']}")
        print(f"克隆函数数量: {result['Clone Func Count']}")
        print(f"函数总数量: {result['Func Count']}")
        print(f"克隆函数占比: {result['Clone Func Ratio']}")
        print("----------")

    return results

def save_as_xlsx(results, file_path):
    # 检查 results 字典是否为空
    if not results:
        print("结果字典为空，未进行任何处理。")
        return

    # 打开现有的工作簿
    wb = openpyxl.load_workbook(file_path)
    version = next(iter(results.keys())).split('\\')[2]
    if version in wb.sheetnames:
        ws = wb[version]  # 获取对应的工作表
    else:
        print(f"工作表 {version} 不存在，未进行任何处理。")
        return

    # 检查标题行（假设是第一行）是否已存在这两列
    headers = [cell.value for cell in ws[1]]  # 获取第一行的所有标题

    # 初始化要添加的列名
    new_columns = ["Clone Func Count", "Clone Func Ratio"]

    # 检查新列是否存在
    columns_to_add = [col for col in new_columns if col not in headers]

    if columns_to_add:
        max_col = ws.max_column
        for col_name in columns_to_add:
            max_col += 1  # 更新最大列数
            ws.cell(row=1, column=max_col, value=col_name)  # 添加列名
            for row in range(2, ws.max_row + 1):  # 从第二行开始填充
                ws.cell(row=row, column=max_col, value=0)

    # 获取新添加列的列索引
    col_clone_count_index = headers.index("Clone Func Count") + 1 if "Clone Func Count" in headers else len(headers) + 1
    col_clone_ratio_index = headers.index("Clone Func Ratio") + 1 if "Clone Func Ratio" in headers else len(headers) + 2

    # 遍历字典结果
    for file_name, result in results.items():
        # 获取表格中已有的路径列，假设路径在第一列
        existing_paths = [cell.value for cell in ws['A'] if cell.value is not None]

        # 直接使用字典中的路径进行匹配
        path = result['File']

        # 如果表格中已经有这个路径，更新对应的行
        if path in existing_paths:
            # 找到行号并更新该行的其他数据
            row_index = existing_paths.index(path) + 1  # Excel 行是从 1 开始的
            ws.cell(row=row_index, column=col_clone_count_index, value=result['Clone Func Count'])
            ws.cell(row=row_index, column=col_clone_ratio_index, value=result['Clone Func Ratio'])

    # 保存修改后的工作簿
    wb.save(file_path)
    print(f"已处理完: {version}")


if __name__ == "__main__":
    xml_root_path = '../clone_xml/autoware_func_cc'
    xml_files = glob.glob(os.path.join(xml_root_path, '*.xml'))
    output_file = '../Results/autoware_result.xlsx'
    for xml_file in xml_files:
        results = {}
        results = calculate(xml_file)
        save_as_xlsx(results, output_file)


# # 现有的 CSV 文件路径
# existing_csv_path = '../Results/autoware_results.csv'
#
# # 读取现有的 CSV 文件
# if os.path.exists(existing_csv_path):
#     existing_df = pd.read_csv(existing_csv_path)
# else:
#     # 如果文件不存在，可以选择创建一个空的 DataFrame
#     existing_df = pd.DataFrame()
#
# # 确保现有 DataFrame 中有 'File Name' 列
# if 'File' not in existing_df.columns:
#     raise ValueError("'File Name' 列在现有 CSV 文件中不存在！")
#
# # 创建新列，初始化为0
# existing_df['Clone Func Count'] = 0
# existing_df['Clone Func Ratio'] = 0.0
#
# # 对现有的 DataFrame 进行遍历，匹配 File Name 并更新相关列
# for idx, row in existing_df.iterrows():
#     file_name = row['File']
#     # 如果在 results 中找到匹配的文件名，则更新相关列
#     if file_name in results:
#         existing_df.at[idx, 'Clone Func Count'] = results[file_name]['Clone Func Count']
#         existing_df.at[idx, 'Clone Func Ratio'] = results[file_name]['Clone Func Ratio']
#
# # 保存更新后的 DataFrame 到 CSV，避免覆盖原文件，使用 mode='a' 来追加
# existing_df.to_csv(existing_csv_path, index=False)
#
# print(f"结果已成功更新到 {existing_csv_path}")
