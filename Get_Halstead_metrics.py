import pandas as pd
import os
from openpyxl import load_workbook

def combine(directory, excel_path):
    # 读取基础 Excel 文件中的所有工作表名称
    base_df = pd.ExcelFile(excel_path)
    base_sheets = base_df.sheet_names

    # 遍历目录中的所有 Excel 文件
    for file_name in os.listdir(directory):
        if file_name.endswith('.xlsx'):
            file_path = os.path.join(directory, file_name)
            version = file_name.split('_')[-1].split('.xlsx')[0]
            print(f"正在处理：{version}的halstead度量信息")

            # 检查基础文件中是否存在对应的工作表
            if version in base_sheets:
                existing_df = pd.read_excel(excel_path, sheet_name=version)
                all_data_df = pd.read_excel(file_path)

                # 根据 'File' 列进行合并，保留原文件的所有列，并将新的度量信息添加到原表中
                merged_df = pd.merge(existing_df, all_data_df, on='File', how='left')

                # 加载工作簿
                wb = load_workbook(excel_path)
                ws = wb[version]

                # 清空原工作表内容（不包括表头）
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.value = None

                # 写入列名
                for col_idx, col_name in enumerate(merged_df.columns, start=1):
                    ws.cell(row=1, column=col_idx, value=col_name)

                # 写入合并后的数据
                for r_idx, row in enumerate(merged_df.values, start=2):
                    for c_idx, value in enumerate(row, start=1):
                        ws.cell(row=r_idx, column=c_idx, value=value)

                # 保存修改后的工作簿
                wb.save(excel_path)
            else:
                print(f"Warning: No sheet named '{version}' found in base file.")

def main():
    directory = input("please input your root directory: ")
    excel_path = input("please input your output_file directory: ")
    # directory = 'Halstead_parser_code/autoware'  # 存放 Excel 文件的目录路径
    # excel_path = 'Results/autoware_result.xlsx'  # 基础 Excel 文件路径
    combine(directory, excel_path)

