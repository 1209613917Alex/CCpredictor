import xml.etree.ElementTree as ET
import pandas as pd
import os
import chardet
from function_parser_code.tree_sitter_parser.tree_sitter_parser import TreeSitterParser
import glob
import openpyxl

from collections import defaultdict

def parse_simian_xml(file_path):
    tree = ET.parse(file_path)
    root = tree.getroot()
    file_to_clones = defaultdict(lambda: defaultdict(int))

    # 存储每个文件的克隆代码块
    file_clone_blocks = defaultdict(list)

    for set_element in root.findall('.//set'):
        for block_element in set_element.findall('.//block'):
            source_file = block_element.get('sourceFile')
            # 获取文件名和文件夹部分
            file_name_only = os.path.basename(source_file)  # 获取文件名，包含扩展名
            file_path = os.path.dirname(source_file)  # 获取路径，去掉文件名
            path_parts = file_path.split(os.sep)
            new_path_parts = path_parts[4:]
            new_path = os.sep.join(new_path_parts)
            if file_name_only.endswith('.cpp'):
                file_name_only = file_name_only[:-4] + '.cc'
            final_path = os.path.join(new_path, file_name_only)

            start_line = int(block_element.get('startLineNumber'))
            end_line = int(block_element.get('endLineNumber'))
            file_clone_blocks[final_path].append((start_line, end_line))

    # 计算每个文件的克隆代码行数和克隆块数
    for file_name, blocks in file_clone_blocks.items():
        sorted_blocks = sorted(blocks, key=lambda x: x[0])  # 按起始行排序
        merged_blocks = []

        for start, end in sorted_blocks:
            if not merged_blocks or start > merged_blocks[-1][1]:
                merged_blocks.append([start, end])
                file_to_clones[file_name]['clone_blocks'] += 1  # 增加克隆块计数
            else:
                merged_blocks[-1] = [merged_blocks[-1][0], max(merged_blocks[-1][1], end)]

        # 计算总的克隆代码行数
        total_clone_lines = sum((end - start + 1) for start, end in merged_blocks)
        file_to_clones[file_name]['clone_lines'] = total_clone_lines

    return file_to_clones

def update_excel_with_clone_counts(excel_path, sheet_name, clone_data):
    try:
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
    except FileNotFoundError:
        print(f"The file {excel_path} does not exist. Creating a new one.")
        df = pd.DataFrame(columns=['File', 'clone_lines', 'clone_blocks', 'clone_lines_ratio', 'CountLine'])

    if 'clone_lines' not in df.columns:
        df['clone_lines'] = 0
    if 'clone_blocks' not in df.columns:
        df['clone_blocks'] = 0
    if 'clone_lines_ratio' not in df.columns:
        df['clone_lines_ratio'] = 0

    updated = False
    for index, row in df.iterrows():
        source_file = row['File'].split(os.sep)
        source_file = os.sep.join(source_file[4:])
        file_clone_data = clone_data.get(source_file)
        if file_clone_data:
            clone_lines = file_clone_data.get('clone_lines', 0)
            clone_blocks = file_clone_data.get('clone_blocks', 0)
            lines = row['CountLine']  # 从Excel表格中获取总行数
            clone_lines_ratio = round(clone_lines / lines, 2) if lines > 0 else 0
            df.at[index, 'clone_lines'] = clone_lines
            df.at[index, 'clone_blocks'] = clone_blocks
            df.at[index, 'clone_lines_ratio'] = clone_lines_ratio
            updated = True

    if updated:
        try:
            with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"Successfully updated sheet {sheet_name} in {excel_path}")
        except Exception as e:
            print(f"Error saving Excel file for {sheet_name}: {e}")
    else:
        print(f"No updates needed for sheet {sheet_name}.")

def func_parse(codes):
    """解析代码提取函数信息"""
    tree_sitter_parser = TreeSitterParser()
    tree_sitter_parser.functions = []  # 清空已存在的函数列表
    tree_sitter_parser.tree_parser(codes)  # 解析代码
    return tree_sitter_parser.functions  # 返回解析出的函数信息

def read_cpp(cpp_adr):
    """读取cpp文件，返回读取结果，如果文件无法打开则返回 None"""
    try:
        # 首先以二进制模式读取文件
        with open(cpp_adr, 'rb') as f:
            data = f.read()

        # 检测文件的编码格式
        detected = chardet.detect(data)
        encode_str = detected['encoding']

        # 读取文件内容，尝试使用检测到的编码
        try:
            with open(cpp_adr, 'r', encoding=encode_str) as f:
                codes = f.read()
            return codes
        except (UnicodeDecodeError, TypeError):
            # 如果解码失败，尝试使用 utf-8 和 latin1
            try:
                with open(cpp_adr, 'r', encoding='utf-8') as f:
                    codes = f.read()
                return codes
            except UnicodeDecodeError:
                # 如果 utf-8 也失败，最后尝试使用 latin1
                with open(cpp_adr, 'r', encoding='latin1') as f:
                    codes = f.read()
                return codes
    except Exception as e:
        print(f"Error opening file {cpp_adr}: {e}")
        return None  # 如果文件无法打开，返回 None

def file_func_num(sourceFile):
    """获取文件中函数的数量"""
    codes = read_cpp(sourceFile)  # 读取cpp文件，获取代码进行函数解析
    functions = func_parse(codes)  # 解析cpp文件中的所有函数信息
    return len(functions)

def calculate(xml_file_path):
    # 解析 XML 文件
    tree = ET.parse(xml_file_path)
    root = tree.getroot()
    # 存储提取的数据
    data = []
    unique_files = set()

    # 遍历所有的 'dup' 节点
    for dup in root.findall('.//dup'):
        count = dup.get('count')  # 获取 'count' 属性值
        # 遍历 'dup' 下的所有 'source' 节点
        for source in dup.findall('source'):
            # 提取每个 'source' 节点的属性
            source_file = source.get('sourceFile')
            unique_files.add(source_file)
            # 提取 function_name 属性
            function_name = source.find('code').get('function_name') if source.find('code') is not None else ""
            # 将提取的信息保存到 data 列表
            data.append({
                'sourceFile': source_file,
                'function_name': function_name
            })

    # 统计每个文件的函数信息
    file_func_data = []
    for unique_file in unique_files:
        unique_func = set()
        file_func_count = file_func_num(unique_file)  # 获取文件中的函数总数
        for entry in data:
            if unique_file == entry['sourceFile']:
                unique_func.add(entry['function_name'])  # 添加克隆函数
        # 存储每个文件的函数数量和克隆函数数量
        file_func_data.append({
            'file_name': unique_file,
            'clone_func_count': len(unique_func),
            'func_count': file_func_count
        })

    # 汇总信息的字典
    results = {}

    # 计算每个文件的克隆函数对的占比（ratio）
    for file in file_func_data:
        file_name = file['file_name']
        clone_count = file['clone_func_count']
        func_count = file['func_count']
        ratio = round(clone_count / func_count,2) if func_count > 0 else 0
        # 将结果添加到汇总字典
        results[file_name] = {
            'File': file_name,
            'Clone Func Count': clone_count,
            'Func Count': func_count,
            'Clone Func Ratio': ratio
        }
    for file_name, result in results.items():
        print(f"文件名: {result['File']}")
        print(f"克隆函数数量: {result['Clone Func Count']}")
        print(f"函数总数量: {result['Func Count']}")
        print(f"克隆函数占比: {result['Clone Func Ratio']}")
        print("----------")

    return results

def save_as_xlsx(results, file_path):
    # 检查 results 字典是否为空
    if not results:
        print("结果字典为空，未进行任何处理。")
        return

    # 打开现有的工作簿
    wb = openpyxl.load_workbook(file_path)
    version = next(iter(results.keys())).split('\\')[2]
    if version in wb.sheetnames:
        ws = wb[version]  # 获 取对应的工作表
    else:
        print(f"工作表 {version} 不存在，未进行任何处理。")
        return

    # 检查标题行（假设是第一行）是否已存在这两列
    headers = [cell.value for cell in ws[1]]  # 获取第一行的所有标题

    # 初始化要添加的列名
    new_columns = ["Clone Func Count", "Clone Func Ratio"]

    # 检查新列是否存在
    columns_to_add = [col for col in new_columns if col not in headers]

    if columns_to_add:
        max_col = ws.max_column
        for col_name in columns_to_add:
            max_col += 1  # 更新最大列数
            ws.cell(row=1, column=max_col, value=col_name)  # 添加列名
            for row in range(2, ws.max_row + 1):  # 从第二行开始填充
                ws.cell(row=row, column=max_col, value=0)

    # 获取新添加列的列索引
    col_clone_count_index = headers.index("Clone Func Count") + 1 if "Clone Func Count" in headers else len(headers) + 1
    col_clone_ratio_index = headers.index("Clone Func Ratio") + 1 if "Clone Func Ratio" in headers else len(headers) + 2

    # 遍历字典结果
    for file_name, result in results.items():
        # 获取表格中已有的路径列，假设路径在第一列
        existing_paths = [cell.value for cell in ws['A'] if cell.value is not None]

        # 直接使用字典中的路径进行匹配
        path = result['File']

        # 如果表格中已经有这个路径，更新对应的行
        if path in existing_paths:
            # 找到行号并更新该行的其他数据
            row_index = existing_paths.index(path) + 1  # Excel 行是从 1 开始的
            ws.cell(row=row_index, column=col_clone_count_index, value=result['Clone Func Count'])
            ws.cell(row=row_index, column=col_clone_ratio_index, value=result['Clone Func Ratio'])

    # 保存修改后的工作簿
    wb.save(file_path)
    print(f"已处理完: {version}")

def clone_func_parser(xml_root_path, output_file):
    xml_files = glob.glob(os.path.join(xml_root_path, '*.xml'))
    for xml_file in xml_files:
        results = calculate(xml_file)
        save_as_xlsx(results, output_file)

def main():
    directory = input("please input your clone_xml directory: ")
    output_path = input("please input your output_file directory: ")
    # example input:
    # directory = 'clone_xml/autoware'
    # output_path = 'Results/autoware_result.xlsx'
    clone_data = {}
    versions  = pd.ExcelFile(output_path).sheet_names
    for version in versions:
        sheet_name = version

        for file_name in os.listdir(directory):
            if file_name.startswith(version) and file_name.endswith('-dup.xml'):
                file_path = os.path.join(directory, file_name)
                clone_counts = parse_simian_xml(file_path)
                for file, counts in clone_counts.items():
                    clone_data[file] = counts

        update_excel_with_clone_counts(output_path, sheet_name, clone_data)
        clone_data.clear()  # 清空clone_data，为下一个版本做准备

    xml_root_path = input("please input your clone_function_xml directory: ")
    output_file = input("please input your output_file directory: ")
    # example input:
    # xml_root_path = 'clone_xml/autoware_func_cc'
    # output_file = 'Results/autoware_result.xlsx'
    clone_func_parser(xml_root_path, output_file)

